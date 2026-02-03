from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
import pyodbc
import pandas as pd
from datetime import datetime
import io
import os
import sys
import time
import threading
import logging
import socket
import urllib.request
import webbrowser
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------
# Path helpers (frozen vs dev)
# -----------------------
def _resource_root() -> str:
    """Resolve resource root for templates/static."""
    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)
        internal_dir = os.path.join(exe_dir, "_internal")

        nested_internal_dir = os.path.join(internal_dir, "_internal")
        if os.path.isdir(os.path.join(nested_internal_dir, "templates")):
            return nested_internal_dir

        if os.path.isdir(os.path.join(internal_dir, "templates")):
            return internal_dir
        if os.path.isdir(os.path.join(exe_dir, "templates")):
            return exe_dir
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass and os.path.isdir(os.path.join(meipass, "templates")):
            return meipass
        return exe_dir
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = _resource_root()

APP_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
try:
    os.chdir(APP_DIR)
except Exception:
    pass

TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)

# -----------------------
# Logging
# -----------------------
LOG_PATH = os.path.join(APP_DIR, "ProductionReportSystem.log")

# 自定義 LOG 過濾器，過濾高頻率請求
class HighFrequencyRequestFilter(logging.Filter):
    """過濾高頻率的心跳和狀態查詢請求"""
    def filter(self, record):
        # 過濾掉以下路徑的 LOG
        filtered_paths = [
            'POST /api/heartbeat',
            'GET /api/get_queue_status',
            'GET /api/get_queue_types'  # 添加此行
        ]
        
        # 檢查訊息中是否包含要過濾的路徑
        for path in filtered_paths:
            if path in record.getMessage():
                return False  # 不記錄此 LOG
        
        return True  # 記錄其他所有 LOG

# 配置 LOG
file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
file_handler.addFilter(HighFrequencyRequestFilter())

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[file_handler],
)
logger = logging.getLogger("prs")

# -----------------------
# Settings
# -----------------------
HOST = "127.0.0.1"
PORT = 5000

_IDLE_TIMEOUT_SEC = 60
_HEARTBEAT_INTERVAL_SEC = 3

_last_heartbeat_ts = time.time()

# 本機暫存目錄（用於存放待上傳的檔案）
LOCAL_EXPORT_DIR = os.path.join(APP_DIR, "exports")
os.makedirs(LOCAL_EXPORT_DIR, exist_ok=True)

# 網路共用資料夾路徑
NETWORK_SHARE_PATH = r"\\sambasy\public\ProductionReportSystem"

# -----------------------
# DB config
# -----------------------
DB_CONFIG = {
    "server": "192.168.0.18",
    "database": "Oee_SingRong_ChiangsTest",
    "username": "jack",
    "password": "730320",
    "driver": "{ODBC Driver 17 for SQL Server}",
}

def get_db_connection():
    """建立資料庫連線"""
    conn_str = (
        f"DRIVER={DB_CONFIG['driver']};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"UID={DB_CONFIG['username']};"
        f"PWD={DB_CONFIG['password']}"
    )
    return pyodbc.connect(conn_str)

def query_production_report(dy_serial_num: str):
    """查詢生產日報表資料"""
    sql = r"""
    SELECT 
        c.DySerialNum AS [生產日報表序號],
        CONVERT(varchar(10), c.CDate, 23) AS [工作日期],
        a.WorkerNum AS [工作者編號],
        a.WorkerName AS [工作者名稱],
        b.ProdNum AS [工序編號],
        b.description AS [工序內容],
        LTRIM(RTRIM(ISNULL(NULLIF(e.PDNum, ''), ''))) AS [發工單號],
        b.PDSerialNum AS [製令序號],
        LTRIM(RTRIM(ISNULL(NULLIF(e.ProdNum, ''), ''))) AS [產品編號],
        LTRIM(RTRIM(ISNULL(NULLIF(e.description, ''), ''))) AS [品名規格],
        a.StartDate AS [起工時間],
        a.FinishDate AS [完工時間],
        CASE 
            WHEN ISNULL(a.csj, 0) = 0 THEN N'標準起工' 
            ELSE N'試模' 
        END AS [起工型態],
        a.MachineNr AS [機台編號],
        TRIM(f.PordDept) AS [機台部門],
        CAST(b.TrueHr AS decimal(18,10)) AS [實際工時],
        b.FinishQty AS [完工數],
        b.BadQty AS [不良數],
        b.ExtraName1 AS [除外名稱1],
        b.OtherHours1 AS [除外時間1],
        b.ExtraName2 AS [除外名稱2],
        b.OtherHours2 AS [除外時間2],
        b.ExtraName3 AS [除外名稱3],
        b.OtherHours3 AS [除外時間3],
        c.EditTime AS [編輯時間]
    FROM dbo.TimeWorkBase a
    LEFT JOIN dbo.DayWorkDYProduct b 
        ON a.DySerialNum = b.DySerialNum 
        AND a.PDSerialNum = b.PDSerialNum 
        AND a.OrdinalNum = b.OrdinalNum
    LEFT JOIN dbo.DayWorkDYBase c 
        ON c.DySerialNum = b.DySerialNum
    LEFT JOIN dbo.ProcessPDBase e 
        ON a.PDSerialNum = e.SerialNum
    LEFT JOIN dbo.Jang1Base f 
        ON a.MachineNr = f.customernr
    WHERE c.DySerialNum = ?
    ORDER BY 
        COALESCE(TRIM(f.PordDept), N'') ASC,
        a.WorkerNum ASC,
        a.StartDate ASC
    """
    try:
        conn = get_db_connection()
        df = pd.read_sql(sql, conn, params=[dy_serial_num])
        conn.close()
        return df
    except Exception as e:
        logger.exception("查詢錯誤: %s", str(e))
        return None


# -----------------------
# 列印清單管理（全域變數，不限筆數）
# -----------------------
print_queue = []  # 每個元素是一個 dict，包含修改資訊
pending_print_records = []  # 臨時存儲待列印的記錄（避免 print_queue 被修改）

# 判斷記錄是否為當天
def is_same_day_record(record: dict) -> bool:
    """判斷記錄的工作日期是否為當天"""
    try:
        # 嘗試多個可能的欄位名稱
        work_date_str = (
            record.get('work_date_original') or 
            record.get('work_date_modified') or 
            record.get('work_date') or 
            ''
        )
        
        logger.info(f"[日期判斷] 序號: {record.get('dy_serial_num')}, work_date_str: {work_date_str}")
        
        if not work_date_str or work_date_str == '':
            logger.warning(f"[日期判斷] 序號 {record.get('dy_serial_num')} 沒有工作日期欄位")
            return False
        
        # 獲取今天的日期（台灣時區）
        today = datetime.now().date()
        logger.info(f"[日期判斷] 今天日期: {today}")
        
        # 轉換為字串並處理
        work_date_str = str(work_date_str).strip()
        
        # 移除時間部分，只保留日期
        if ' ' in work_date_str:
            date_part = work_date_str.split(' ')[0]
        else:
            date_part = work_date_str
        
        # 替換 / 為 -
        date_part = date_part.replace('/', '-')
        
        logger.info(f"[日期判斷] 處理後的日期字串: {date_part}")
        
        # 解析日期
        try:
            work_date = datetime.strptime(date_part, '%Y-%m-%d').date()
            result = work_date == today
            logger.info(f"[日期判斷] 工作日期: {work_date}, 是否為當天: {result}")
            return result
        except ValueError as ve:
            logger.error(f"[日期判斷] 日期格式錯誤: {date_part}, 錯誤: {str(ve)}")
            return False
        
    except Exception as e:
        logger.exception(f"[日期判斷] 判斷失敗，序號: {record.get('dy_serial_num')}, 錯誤: {str(e)}")
        return False

# -----------------------
# Excel 套表生成函數
# -----------------------
def create_print_template(records: list) -> io.BytesIO:
    """
    創建列印套表（新版本：一張Excel包含2張修改申請）
    
    重要：
    - 只使用 A-F 欄（6欄）
    - 縮放比例 56%
    - 2 張修改申請左右排列（不再有下方的表格）
    - 合併儲存格的資料要放在起始儲存格（左上角）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "生產日報表"
    
    # 設定頁面
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # 頁面邊界（根據範本）
    ws.page_margins.left = 0.0
    ws.page_margins.right = 0.0
    ws.page_margins.top = 0.0
    ws.page_margins.bottom = 0.3937007874015748  # 1公分
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    
    # 置中方式：水平置中和垂直置中
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    # 關鍵：設定縮放比例為 56%
    ws.page_setup.scale = 56
    
    # 符合頁數設定
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # 設定頁尾
    ws.oddFooter.right.text = "生管：__________"
    ws.oddFooter.right.font = "新細明體,粗體"
    ws.oddFooter.right.size = 36
    
    # 樣式定義（正確的字體大小）
    title_font = Font(name='新細明體', size=90, bold=True)
    normal_font = Font(name='新細明體', size=30)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 標題（A1:F3 合併）
    ws.merge_cells('A1:F3')
    title_cell = ws['A1']
    title_cell.value = "生產日報表 修改申請"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 41.25
    ws.row_dimensions[2].height = 41.25
    ws.row_dimensions[3].height = 41.25
    
    # 定義欄位和其數值格式
    fields = [
        ('工作日期', 'work_date', 'mm-dd-yy'),
        ('工作者編號', 'worker_num', 'General'),
        ('機台代號', 'machine_num', 'General'),
        ('工序編號', 'prod_num', 'General'),
        ('完工數', 'finish_qty', 'General'),
        ('不良數', 'bad_qty', 'General'),
        ('起工時間', 'start_time', 'yyyy/m/d h:mm'),
        ('完工時間', 'finish_time', 'yyyy/m/d h:mm'),
        ('除外名稱1', 'extra_name1', 'General'),
        ('除外時間1', 'extra_time1', 'General'),
        ('除外名稱2', 'extra_name2', 'General'),
        ('除外時間2', 'extra_time2', 'General'),
        ('除外名稱3', 'extra_name3', 'General'),
        ('除外時間3', 'extra_time3', 'General'),
    ]
    
    # 2 個表格的配置（左右排列）
    configs = [
        # 左邊
        {'start_row': 4, 'label_col': 'A', 'value_col': 'B', 'mod_col': 'C',
         'delete_label': '生產日報表刪除', 'delete_value_col': 'C'},
        # 右邊
        {'start_row': 4, 'label_col': 'D', 'value_col': 'E', 'mod_col': 'F',
         'delete_label': '生產日報表刪除', 'delete_value_col': 'F'},
    ]
    
    # 固定畫出 2 個表格（左右排列）
    for idx, config in enumerate(configs):
        start_row = config['start_row']
        col_label = config['label_col']
        col_value = config['value_col']
        col_mod = config['mod_col']
        delete_label = config['delete_label']
        delete_value_col = config['delete_value_col']
        
        # 取得資料
        rec = records[idx] if idx < len(records) else {}
        
        current_row = start_row
        
        # 第 1 行：★生產日報表序號（value:mod 合併）
        cell = ws[f'{col_label}{current_row}']
        cell.value = "★生產日報表序號"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # 合併 value:mod 欄，資料放在 value 欄（合併起始儲存格）
        ws.merge_cells(f'{col_value}{current_row}:{col_mod}{current_row}')
        cell = ws[f'{col_value}{current_row}']
        cell.value = rec.get('dy_serial_num', '')
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # 也要設定 mod 欄的框線
        cell = ws[f'{col_mod}{current_row}']
        cell.border = thin_border
        ws.row_dimensions[current_row].height = 39.95
        
        current_row += 1
        
        # 第 2 行：發工單號（value:mod 合併）
        cell = ws[f'{col_label}{current_row}']
        cell.value = "發工單號"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # 合併 value:mod 欄，資料放在 value 欄（合併起始儲存格）
        ws.merge_cells(f'{col_value}{current_row}:{col_mod}{current_row}')
        cell = ws[f'{col_value}{current_row}']
        cell.value = rec.get('pd_num', '')
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # 也要設定 mod 欄的框線
        cell = ws[f'{col_mod}{current_row}']
        cell.border = thin_border
        ws.row_dimensions[current_row].height = 39.95
        
        current_row += 1
        
        # 第 3 行：刪除行（不合併，各自獨立）
        cell = ws[f'{col_label}{current_row}']
        cell.value = delete_label
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # 刪除值的處理
        delete_value = rec.get('delete_flag', '')
        if delete_value == '是':
            delete_mark = 'Y'
        elif delete_value == '否':
            delete_mark = 'N'
        else:
            delete_mark = ''
        
        # 原本欄位（col_value）
        cell = ws[f'{col_value}{current_row}']
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # 根據 delete_value_col 決定是否要在這裡顯示刪除標記
        if delete_value_col == col_value:
            cell.value = delete_mark
        
        # 修改為欄位（col_mod）
        cell = ws[f'{col_mod}{current_row}']
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # 根據 delete_value_col 決定是否要在這裡顯示刪除標記
        if delete_value_col == col_mod:
            cell.value = delete_mark
        
        ws.row_dimensions[current_row].height = 39.95
        
        current_row += 1
        
        # 第 4 行：原本/修改為標題（不合併）
        cell = ws[f'{col_value}{current_row}']
        cell.value = "原本"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        cell = ws[f'{col_mod}{current_row}']
        cell.value = "修改為"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        ws.row_dimensions[current_row].height = 39.95
        
        current_row += 1
        
        # 14 個欄位（全部不合併）
        for field_name, field_key, number_format in fields:
            cell = ws[f'{col_label}{current_row}']
            cell.value = field_name
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # 處理原本值
            cell = ws[f'{col_value}{current_row}']
            
            # 如果是刪除操作，原本欄位應該留空
            if rec.get('delete_flag') == '是':
                original_value = ''
            else:
                original_value = rec.get(f'{field_key}_original', '')
                # 如果是日期時間欄位，轉換字串為 datetime 物件
                if field_key in ['start_time', 'finish_time'] and original_value:
                    try:
                        if isinstance(original_value, str):
                            # 嘗試解析不同格式的日期時間字串
                            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M']:
                                try:
                                    original_value = datetime.strptime(original_value, fmt)
                                    break
                                except ValueError:
                                    continue
                    except Exception:
                        pass  # 如果轉換失敗，保持原值
            
            cell.value = original_value
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.number_format = number_format
            
            # 處理修改值
            cell = ws[f'{col_mod}{current_row}']
            modified_value = rec.get(f'{field_key}_modified', '')
            # 如果是日期時間欄位，轉換字串為 datetime 物件
            if field_key in ['start_time', 'finish_time'] and modified_value:
                try:
                    if isinstance(modified_value, str):
                        # 嘗試解析不同格式的日期時間字串
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M']:
                            try:
                                modified_value = datetime.strptime(modified_value, fmt)
                                break
                            except ValueError:
                                continue
                except Exception:
                    pass  # 如果轉換失敗，保持原值
            cell.value = modified_value
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.number_format = number_format
            
            ws.row_dimensions[current_row].height = 39.95
            current_row += 1
    
    # 底部合併儲存格（第 22-23 行合併，但內容為空）
    ws.merge_cells('A22:F23')
    # 不設定值，保持為空
    ws.row_dimensions[22].height = 41.25
    ws.row_dimensions[23].height = 41.25
    
    # 設定欄寬（按照範本）
    ws.column_dimensions['A'].width = 57.28515625
    ws.column_dimensions['B'].width = 40.7109375
    ws.column_dimensions['C'].width = 40.0  # 修正為 40
    ws.column_dimensions['D'].width = 57.28515625
    ws.column_dimensions['E'].width = 40.7109375
    ws.column_dimensions['F'].width = 40.0  # 修正為 40
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_multiple_excel_files(records: list) -> list:
    """
    根據記錄數量生成多個 Excel 檔案（每 2 筆一個檔案）
    **只為非當天記錄生成 Excel**（當天記錄只需要 CSV）
    返回生成的檔案路徑列表
    """
    if not records:
        return []
    
    # 只為非當天記錄生成 Excel
    different_day_records = [r for r in records if not is_same_day_record(r)]
    
    if not different_day_records:
        logger.info("沒有非當天記錄，不生成 Excel")
        return []
    
    logger.info(f"準備為 {len(different_day_records)} 筆非當天記錄生成 Excel")
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_files = []
    
    # 每 2 筆記錄生成一個 Excel
    for i in range(0, len(different_day_records), 2):
        batch = different_day_records[i:i+2]  # 取 2 筆（最後可能只有 1 筆）
        batch_num = i // 2 + 1
        
        filename = f"生產日報表修改申請_{len(batch)}筆_批次{batch_num}_{timestamp}.xlsx"
        filepath = os.path.join(LOCAL_EXPORT_DIR, filename)
        
        # 生成 Excel
        excel_output = create_print_template(batch)
        with open(filepath, 'wb') as f:
            f.write(excel_output.getvalue())
        
        excel_files.append(filepath)
        logger.info(f"已生成 Excel 批次 {batch_num}: {filename}（非當天記錄）")
    
    return excel_files


def generate_print_urls(records: list) -> list:
    """
    生成單一列印頁面 URL，包含所有記錄
    返回只有一個 URL 的列表
    """
    if not records:
        return []
    
    # 生成包含所有記錄索引的單一 URL
    all_indices = ','.join(str(i) for i in range(len(records)))
    url = f"/print_page?indices={all_indices}"
    
    return [url]  # 返回只包含一個 URL 的列表


# -----------------------
# CSV 生成函數
# -----------------------
def create_csv_export(record: dict) -> str:
    """生成CSV檔案，返回檔案路徑"""
    dy_serial_num = record.get('dy_serial_num', 'UNKNOWN')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"生產日報表修改_{dy_serial_num}_{timestamp}.csv"
    filepath = os.path.join(LOCAL_EXPORT_DIR, filename)
    
    # CSV 欄位順序
    csv_columns = [
        '生產日報表序號', '刪除(Y/N)', '發工單號', '工作日期', '工作者編號',
        '機台編號', '工序編號', '完工數', '不良數', '起工時間', '完工時間',
        '除外名稱1', '除外時間1', '除外名稱2', '除外時間2', '除外名稱3', '除外時間3',
        '儲存時間'
    ]
    
    # 轉換刪除標記格式：是→Y，否→N
    delete_flag = record.get('delete_flag', '否')
    if delete_flag == '是':
        delete_flag_csv = 'Y'
    elif delete_flag == '否':
        delete_flag_csv = 'N'
    else:
        delete_flag_csv = 'N'  # 預設為 N
    
    # 準備資料
    data_row = {
        '生產日報表序號': record.get('dy_serial_num', ''),
        '刪除(Y/N)': delete_flag_csv,
        '發工單號': record.get('pd_num', ''),
        '工作日期': record.get('work_date_modified', record.get('work_date_original', '')),
        '工作者編號': record.get('worker_num_modified', record.get('worker_num_original', '')),
        '機台編號': record.get('machine_num_modified', record.get('machine_num_original', '')),
        '工序編號': record.get('prod_num_modified', record.get('prod_num_original', '')),
        '完工數': record.get('finish_qty_modified', record.get('finish_qty_original', '')),
        '不良數': record.get('bad_qty_modified', record.get('bad_qty_original', '')),
        '起工時間': record.get('start_time_modified', record.get('start_time_original', '')),
        '完工時間': record.get('finish_time_modified', record.get('finish_time_original', '')),
        '除外名稱1': record.get('extra_name1_modified', record.get('extra_name1_original', '')),
        '除外時間1': record.get('extra_time1_modified', record.get('extra_time1_original', '')),
        '除外名稱2': record.get('extra_name2_modified', record.get('extra_name2_original', '')),
        '除外時間2': record.get('extra_time2_modified', record.get('extra_time2_original', '')),
        '除外名稱3': record.get('extra_name3_modified', record.get('extra_name3_original', '')),
        '除外時間3': record.get('extra_time3_modified', record.get('extra_time3_original', '')),
        '儲存時間': record.get('saved_time', ''),
    }
    
    df = pd.DataFrame([data_row], columns=csv_columns)
    df.to_csv(filepath, index=False, encoding='utf-8-sig')
    
    logger.info(f"CSV 已生成: {filepath}")
    return filepath


# -----------------------
# 網路路徑上傳函數
# -----------------------
def upload_to_network_share(local_filepath: str) -> bool:
    """將檔案上傳到網路共用資料夾"""
    try:
        # 檢查網路路徑是否可用
        if not os.path.exists(NETWORK_SHARE_PATH):
            logger.error(f"網路路徑不存在或無法訪問: {NETWORK_SHARE_PATH}")
            return False
        
        filename = os.path.basename(local_filepath)
        dest_path = os.path.join(NETWORK_SHARE_PATH, filename)
        
        shutil.copy2(local_filepath, dest_path)
        logger.info(f"檔案已上傳: {dest_path}")
        return True
        
    except Exception as e:
        logger.exception(f"上傳失敗: {str(e)}")
        return False


# -----------------------
# Web routes
# -----------------------
@app.route("/")
def index():
    return redirect(url_for("table_view"))

@app.route("/table")
def table_view():
    return render_template("index_table.html")

@app.route("/print_page")
def print_page():
    """顯示列印頁面（根據索引顯示 1-2 筆記錄）"""
    global pending_print_records
    
    indices_str = request.args.get('indices', '')
    
    if not indices_str:
        return "缺少記錄索引", 400
    
    try:
        indices = [int(i) for i in indices_str.split(',')]
    except ValueError:
        return "無效的索引格式", 400
    
    # 從 pending_print_records 中取得對應的記錄
    records = []
    for idx in indices:
        if 0 <= idx < len(pending_print_records):
            records.append(pending_print_records[idx])
    
    if not records:
        logger.warning(f"列印頁面找不到記錄，索引：{indices_str}, pending_print_records 數量：{len(pending_print_records)}")
        return "找不到記錄", 404
    
    logger.info(f"列印頁面顯示 {len(records)} 筆記錄")
    
    # 列印頁面顯示後，清空 pending_print_records
    # pending_print_records = []  # 不要立即清空，可能需要重新列印
    
    return render_template("print_template.html", records=records)

@app.route("/health")
def health():
    return "ok"

@app.route("/api/heartbeat", methods=["POST"])
def api_heartbeat():
    global _last_heartbeat_ts
    _last_heartbeat_ts = time.time()
    return jsonify({"success": True})

@app.route("/api/closing", methods=["POST"])
def api_closing():
    global _last_heartbeat_ts
    _last_heartbeat_ts = time.time() - (_IDLE_TIMEOUT_SEC + 5)
    return jsonify({"success": True})

@app.route("/shutdown", methods=["GET", "POST"])
def shutdown():
    if request.remote_addr not in ("127.0.0.1", "::1"):
        return "forbidden", 403

    func = request.environ.get("werkzeug.server.shutdown")
    if func:
        func()
        return "shutting down", 200

    os._exit(0)

@app.route("/api/query", methods=["POST"])
def api_query():
    data = request.get_json() or {}
    dy_serial_num = (data.get("dySerialNum") or "").strip()

    if not dy_serial_num:
        return jsonify({"success": False, "message": "請輸入生產日報表序號"})

    # 如果沒有 DY 前綴，自動添加
    if not dy_serial_num.upper().startswith("DY"):
        dy_serial_num = "DY" + dy_serial_num
    
    # 統一轉為大寫
    dy_serial_num = dy_serial_num.upper()

    df = query_production_report(dy_serial_num)

    if df is None:
        return jsonify({"success": False, "message": "資料庫查詢錯誤"})

    if df.empty:
        return jsonify({"success": False, "message": "查無資料"})

    # datetime -> str
    for col in df.columns:
        if str(df[col].dtype).startswith("datetime"):
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S")

    df = df.where(pd.notnull(df), None)

    return jsonify(
        {
            "success": True,
            "data": df.to_dict("records"),
            "count": len(df),
        }
    )

@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json() or {}
    dy_serial_num = (data.get("dySerialNum") or "").strip()

    if not dy_serial_num:
        return jsonify({"success": False, "message": "請輸入生產日報表序號"})

    # 如果沒有 DY 前綴，自動添加
    if not dy_serial_num.upper().startswith("DY"):
        dy_serial_num = "DY" + dy_serial_num
    
    # 統一轉為大寫
    dy_serial_num = dy_serial_num.upper()

    df = query_production_report(dy_serial_num)

    if df is None or df.empty:
        return jsonify({"success": False, "message": "無資料可匯出"})

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="生產日報表", index=False)

    output.seek(0)
    filename = f"生產日報表_{dy_serial_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/save", methods=["POST"])
def api_save():
    """儲存修改資訊（加入列印清單 + 生成Excel/CSV）"""
    global print_queue
    
    data = request.get_json() or {}
    
    # 驗證：至少要有1個欄位有輸入
    has_modification = False
    editable_fields = [
        'work_date', 'worker_num', 'machine_num', 'prod_num',
        'finish_qty', 'bad_qty', 'start_time', 'finish_time',
        'extra_name1', 'extra_time1', 'extra_name2', 'extra_time2',
        'extra_name3', 'extra_time3'
    ]
    
    delete_flag = data.get('delete_flag', '否')
    
    # 如果勾選刪除，則不需要其他欄位
    if delete_flag == '是':
        has_modification = True
    else:
        for field in editable_fields:
            if data.get(f'{field}_modified', '').strip():
                has_modification = True
                break
    
    if not has_modification:
        return jsonify({"success": False, "message": "尚未輸入任何修改資訊"})
    
    # 加入列印清單（不限筆數）
    # 不再檢查上限
    
    # 記錄儲存時間
    saved_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data['saved_time'] = saved_time
    
    print_queue.append(data)
    
    # 刪除舊的 Excel 檔案（如果存在）
    try:
        for f in os.listdir(LOCAL_EXPORT_DIR):
            if f.startswith('生產日報表修改申請_') and f.endswith('.xlsx'):
                os.remove(os.path.join(LOCAL_EXPORT_DIR, f))
    except Exception as e:
        logger.warning(f"清理舊檔案失敗: {str(e)}")
    
    # 生成 Excel 檔案（每 2 筆一個檔案）
    try:
        excel_files = create_multiple_excel_files(print_queue)
        logger.info(f"已生成 {len(excel_files)} 個 Excel 檔案（共 {len(print_queue)} 筆記錄）")
        excel_filenames = [os.path.basename(f) for f in excel_files]
    except Exception as e:
        logger.exception(f"生成 Excel 失敗: {str(e)}")
        return jsonify({"success": False, "message": f"儲存失敗: {str(e)}"})
    
    # 生成 CSV 檔案（每筆記錄一個 CSV）
    try:
        csv_filepath = create_csv_export(data)
    except Exception as e:
        logger.exception(f"生成 CSV 失敗: {str(e)}")
        return jsonify({"success": False, "message": f"儲存失敗: {str(e)}"})
    
    return jsonify({
        "success": True,
        "message": f"已儲存至列印清單（目前 {len(print_queue)} 筆）",
        "queue_count": len(print_queue),
        "excel_files": excel_filenames,
        "csv_file": os.path.basename(csv_filepath)
    })


@app.route("/api/upload", methods=["POST"])
def api_upload():
    """上傳當天記錄到網路資料夾（CSV + 所有Excel）"""
    global print_queue
    
    try:
        # 分類記錄：當天 vs 非當天
        same_day_records = [r for r in print_queue if is_same_day_record(r)]
        different_day_records = [r for r in print_queue if not is_same_day_record(r)]
        
        if not same_day_records:
            return jsonify({"success": False, "message": "沒有當天記錄可上傳"})
        
        # 取得當天記錄的序號
        same_day_serials = [r.get('dy_serial_num', '') for r in same_day_records]
        logger.info(f"準備上傳當天記錄：{same_day_serials}")
        
        # 找出當天記錄對應的 CSV 檔案
        csv_files_to_upload = []
        for filename in os.listdir(LOCAL_EXPORT_DIR):
            if not filename.endswith('.csv'):
                continue
            
            # 檢查檔案名稱是否包含當天記錄的序號
            for serial in same_day_serials:
                if serial in filename:
                    csv_files_to_upload.append(filename)
                    break
        
        # 找出所有 Excel 檔案（不管當天或非當天）
        excel_files_to_upload = []
        for filename in os.listdir(LOCAL_EXPORT_DIR):
            if filename.startswith('生產日報表修改申請_') and filename.endswith('.xlsx'):
                excel_files_to_upload.append(filename)
        
        files_to_upload = csv_files_to_upload + excel_files_to_upload
        
        if not files_to_upload:
            return jsonify({"success": False, "message": "沒有檔案可上傳"})
        
        logger.info(f"找到當天記錄的 CSV：{csv_files_to_upload}")
        logger.info(f"找到所有 Excel：{excel_files_to_upload}")
        
        # 上傳檔案
        success_count = 0
        failed_files = []
        
        for filename in files_to_upload:
            filepath = os.path.join(LOCAL_EXPORT_DIR, filename)
            if upload_to_network_share(filepath):
                success_count += 1
                logger.info(f"上傳成功：{filename}")
            else:
                failed_files.append(filename)
                logger.error(f"上傳失敗：{filename}")
        
        if failed_files:
            return jsonify({
                "success": False,
                "message": f"部分檔案上傳失敗: {', '.join(failed_files)}"
            })
        
        # 上傳成功後，從修改申請清單中移除當天記錄
        removed_count = len(same_day_records)
        print_queue = different_day_records
        
        logger.info(f"從清單移除 {removed_count} 筆當天記錄，剩餘 {len(print_queue)} 筆非當天記錄")
        
        # 刪除已上傳的本機檔案
        try:
            for filename in files_to_upload:
                filepath = os.path.join(LOCAL_EXPORT_DIR, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    logger.info(f"刪除本機檔案：{filename}")
        except Exception as e:
            logger.warning(f"清理本機檔案失敗: {str(e)}")
        
        # 不需要重新生成 Excel
        # 非當天記錄的 Excel 已經在儲存時生成，不需要重複生成
        if different_day_records:
            logger.info(f"剩餘 {len(different_day_records)} 筆非當天記錄（Excel 已在儲存時生成）")
        
        return jsonify({
            "success": True,
            "message": f"已成功上傳 {removed_count} 筆當天記錄（{len(csv_files_to_upload)} 個 CSV + {len(excel_files_to_upload)} 個 Excel），" +
                      (f"修改申請清單還有 {len(different_day_records)} 筆非當天記錄" if different_day_records else "修改申請清單已清空"),
            "queue_count": len(different_day_records)
        })
        
    except Exception as e:
        logger.exception(f"上傳失敗: {str(e)}")
        return jsonify({"success": False, "message": f"上傳失敗: {str(e)}"})


@app.route("/api/print", methods=["POST"])
def api_print():
    """列印非當天記錄（只處理非當天記錄）"""
    global print_queue
    
    if not print_queue:
        return jsonify({"success": False, "message": "修改申請清單為空"})
    
    try:
        # 分類記錄：當天 vs 非當天
        same_day_records = [r for r in print_queue if is_same_day_record(r)]
        different_day_records = [r for r in print_queue if not is_same_day_record(r)]
        
        if not different_day_records:
            return jsonify({"success": False, "message": "沒有非當天記錄可列印"})
        
        # 取得非當天記錄的序號
        different_day_serials = [r.get('dy_serial_num', '') for r in different_day_records]
        logger.info(f"準備列印非當天記錄：{different_day_serials}")
        
        # 找出非當天記錄對應的 CSV 檔案
        csv_files_to_upload = []
        for filename in os.listdir(LOCAL_EXPORT_DIR):
            if not filename.endswith('.csv'):
                continue
            
            # 檢查檔案名稱是否包含非當天記錄的序號
            for serial in different_day_serials:
                if serial in filename:
                    csv_files_to_upload.append(filename)
                    break
        
        # 找出所有 Excel 檔案
        excel_files_to_upload = []
        for filename in os.listdir(LOCAL_EXPORT_DIR):
            if filename.startswith('生產日報表修改申請_') and filename.endswith('.xlsx'):
                excel_files_to_upload.append(filename)
        
        files_to_upload = csv_files_to_upload + excel_files_to_upload
        
        if not files_to_upload:
            return jsonify({"success": False, "message": "沒有檔案可上傳"})
        
        logger.info(f"找到非當天記錄的 CSV：{csv_files_to_upload}")
        logger.info(f"找到所有 Excel：{excel_files_to_upload}")
        
        # 上傳非當天記錄的檔案
        upload_success_count = 0
        upload_failed_files = []
        
        for filename in files_to_upload:
            filepath = os.path.join(LOCAL_EXPORT_DIR, filename)
            if upload_to_network_share(filepath):
                upload_success_count += 1
                logger.info(f"上傳成功：{filename}")
            else:
                upload_failed_files.append(filename)
                logger.error(f"上傳失敗：{filename}")
        
        if upload_failed_files:
            return jsonify({
                "success": False,
                "message": f"部分檔案上傳失敗: {', '.join(upload_failed_files)}"
            })
        
        # 將非當天記錄存儲到臨時變數，供列印頁面使用
        global pending_print_records
        pending_print_records = different_day_records.copy()
        
        # 生成列印頁面 URL（使用簡單的索引）
        indices_str = ','.join(str(i) for i in range(len(pending_print_records)))
        print_url = f"/print_page?indices={indices_str}"
        print_urls = [print_url]
        
        logger.info(f"已生成列印頁面 URL，待列印記錄數：{len(pending_print_records)}")
        
        # 從修改申請清單中移除非當天記錄
        removed_count = len(different_day_records)
        print_queue = same_day_records
        
        logger.info(f"從清單移除 {removed_count} 筆非當天記錄，剩餘 {len(print_queue)} 筆當天記錄")
        
        # 刪除已上傳的本機檔案
        try:
            for filename in files_to_upload:
                filepath = os.path.join(LOCAL_EXPORT_DIR, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    logger.info(f"刪除本機檔案：{filename}")
        except Exception as e:
            logger.warning(f"清理本機檔案失敗: {str(e)}")
        
        # 不需要重新生成 Excel
        # Excel 已經在儲存時生成，直接使用已上傳的檔案即可
        if same_day_records:
            logger.info(f"剩餘 {len(same_day_records)} 筆當天記錄（不需要生成 Excel）")
        
        return jsonify({
            "success": True,
            "message": f"已上傳 {removed_count} 筆非當天記錄（{len(csv_files_to_upload)} 個 CSV + {len(excel_files_to_upload)} 個 Excel），已生成列印頁面" +
                      (f"；修改申請清單還有 {len(same_day_records)} 筆當天記錄" if same_day_records else "；修改申請清單已清空"),
            "print_urls": print_urls,
            "queue_count": len(same_day_records)
        })
        
    except Exception as e:
        logger.exception(f"列印失敗: {str(e)}")
        return jsonify({"success": False, "message": f"列印失敗: {str(e)}"})


@app.route("/api/get_queue_types", methods=["GET"])
def api_get_queue_types():
    """獲取修改申請清單中當天和非當天記錄的數量"""
    global print_queue
    
    same_day_count = sum(1 for r in print_queue if is_same_day_record(r))
    different_day_count = len(print_queue) - same_day_count
    
    return jsonify({
        "success": True,
        "same_day_count": same_day_count,
        "different_day_count": different_day_count,
        "total_count": len(print_queue)
    })


@app.route("/api/clear_queue", methods=["POST"])
def api_clear_queue():
    """清空所有列印清單"""
    global print_queue
    count = len(print_queue)
    print_queue = []
    
    # 刪除所有生成的 Excel 和 CSV 檔案
    try:
        for f in os.listdir(LOCAL_EXPORT_DIR):
            if (f.startswith('生產日報表修改申請_') and f.endswith('.xlsx')) or \
               (f.startswith('生產日報表修改_') and f.endswith('.csv')):
                filepath = os.path.join(LOCAL_EXPORT_DIR, f)
                os.remove(filepath)
                logger.info(f"已刪除: {f}")
    except Exception as e:
        logger.warning(f"清理檔案失敗: {str(e)}")
    
    return jsonify({
        "success": True,
        "message": f"已清空列印清單（原有 {count} 筆）"
    })


@app.route("/api/clear_same_day_queue", methods=["POST"])
def api_clear_same_day_queue():
    """清空當天修改的記錄（上傳後調用）"""
    global print_queue
    
    # 找出當天的記錄
    same_day_records = [r for r in print_queue if r.get('date_type') == 'same_day']
    same_day_serials = [r.get('dy_serial_num') for r in same_day_records]
    
    # 移除當天的記錄
    print_queue = [r for r in print_queue if r.get('date_type') != 'same_day']
    
    # 刪除當天記錄對應的檔案
    try:
        for f in os.listdir(LOCAL_EXPORT_DIR):
            if not (f.endswith('.xlsx') or f.endswith('.csv')):
                continue
            
            # 檢查檔案是否屬於當天的記錄
            is_same_day_file = any(serial in f for serial in same_day_serials)
            
            if is_same_day_file:
                filepath = os.path.join(LOCAL_EXPORT_DIR, f)
                os.remove(filepath)
                logger.info(f"已刪除當天記錄檔案: {f}")
    except Exception as e:
        logger.warning(f"清理當天記錄檔案失敗: {str(e)}")
    
    return jsonify({
        "success": True,
        "message": f"已清空當天修改記錄（{len(same_day_records)} 筆）"
    })


@app.route("/api/clear_different_day_queue", methods=["POST"])
def api_clear_different_day_queue():
    """清空非當天修改的記錄（列印後調用）"""
    global print_queue
    
    # 找出非當天的記錄
    different_day_records = [r for r in print_queue if r.get('date_type') == 'different_day']
    different_day_serials = [r.get('dy_serial_num') for r in different_day_records]
    
    # 移除非當天的記錄
    print_queue = [r for r in print_queue if r.get('date_type') != 'different_day']
    
    # 刪除非當天記錄對應的檔案
    try:
        for f in os.listdir(LOCAL_EXPORT_DIR):
            if not (f.endswith('.xlsx') or f.endswith('.csv')):
                continue
            
            # 檢查檔案是否屬於非當天的記錄
            is_different_day_file = any(serial in f for serial in different_day_serials)
            
            if is_different_day_file:
                filepath = os.path.join(LOCAL_EXPORT_DIR, f)
                os.remove(filepath)
                logger.info(f"已刪除非當天記錄檔案: {f}")
    except Exception as e:
        logger.warning(f"清理非當天記錄檔案失敗: {str(e)}")
    
    return jsonify({
        "success": True,
        "message": f"已清空非當天修改記錄（{len(different_day_records)} 筆）"
    })


@app.route("/api/get_queue_status", methods=["GET"])
def api_get_queue_status():
    """取得列印清單狀態"""
    return jsonify({
        "success": True,
        "queue_count": len(print_queue),
        "queue": print_queue
    })


@app.route("/api/delete_queue_item", methods=["POST"])
def api_delete_queue_item():
    """刪除列印清單中的單一筆記錄"""
    global print_queue
    
    data = request.get_json() or {}
    index = data.get('index')
    
    if index is None or index < 0 or index >= len(print_queue):
        return jsonify({"success": False, "message": "無效的索引"})
    
    deleted_item = print_queue.pop(index)
    deleted_serial_num = deleted_item.get('dy_serial_num', 'UNKNOWN')
    
    # 刪除對應的 CSV 檔案
    try:
        for f in os.listdir(LOCAL_EXPORT_DIR):
            # 檢查是否為該序號的 CSV 檔案
            if f.startswith(f'生產日報表修改_{deleted_serial_num}_') and f.endswith('.csv'):
                csv_filepath = os.path.join(LOCAL_EXPORT_DIR, f)
                os.remove(csv_filepath)
                logger.info(f"已刪除 CSV: {f}")
                break
    except Exception as e:
        logger.warning(f"刪除 CSV 失敗: {str(e)}")
    
    # 重新生成 Excel（包含剩餘的記錄）
    if print_queue:
        try:
            # 刪除舊的 Excel 檔案
            for f in os.listdir(LOCAL_EXPORT_DIR):
                if f.startswith('生產日報表修改申請_') and f.endswith('.xlsx'):
                    os.remove(os.path.join(LOCAL_EXPORT_DIR, f))
            
            # 生成新的 Excel（可能有多個檔案）
            excel_files = create_multiple_excel_files(print_queue)
            
            logger.info(f"已刪除記錄並重新生成 {len(excel_files)} 個 Excel 檔案")
        except Exception as e:
            logger.exception(f"重新生成 Excel 失敗: {str(e)}")
    else:
        # 如果清單為空，刪除所有 Excel 檔案
        try:
            for f in os.listdir(LOCAL_EXPORT_DIR):
                if f.startswith('生產日報表修改申請_') and f.endswith('.xlsx'):
                    os.remove(os.path.join(LOCAL_EXPORT_DIR, f))
        except Exception as e:
            logger.warning(f"清理檔案失敗: {str(e)}")
    
    return jsonify({
        "success": True,
        "message": f"已刪除記錄（生產日報表序號：{deleted_serial_num}）",
        "queue_count": len(print_queue)
    })


# -----------------------
# Process control
# -----------------------
def _is_our_server_running() -> bool:
    """Check if something is listening on HOST:PORT and it responds /health = ok."""
    try:
        with urllib.request.urlopen(f"http://{HOST}:{PORT}/health", timeout=0.6) as resp:
            body = resp.read(32).decode("utf-8", errors="ignore")
            return body.strip() == "ok"
    except Exception:
        return False

def _open_browser():
    try:
        webbrowser.open(f"http://{HOST}:{PORT}/table", new=1, autoraise=True)
    except Exception as e:
        logger.exception("無法開啟瀏覽器: %s", str(e))

def _request_shutdown():
    try:
        urllib.request.urlopen(f"http://{HOST}:{PORT}/shutdown", timeout=1)
    except Exception:
        pass

def _idle_monitor():
    """If no heartbeat for IDLE seconds, shut down server."""
    global _last_heartbeat_ts
    while True:
        time.sleep(2)
        idle = time.time() - _last_heartbeat_ts
        if idle > _IDLE_TIMEOUT_SEC:
            logger.info("Idle %ss > %ss. Shutting down.", int(idle), _IDLE_TIMEOUT_SEC)
            _request_shutdown()
            time.sleep(2)
            os._exit(0)

def main():
    if _is_our_server_running():
        _open_browser()
        return

    t = threading.Thread(target=_idle_monitor, daemon=True)
    t.start()

    threading.Timer(1.0, _open_browser).start()

    logger.info("Starting server at http://%s:%s", HOST, PORT)

    app.run(host=HOST, port=PORT, debug=False, use_reloader=False, threaded=True)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.exception("Fatal error: %s", str(e))
        raise
